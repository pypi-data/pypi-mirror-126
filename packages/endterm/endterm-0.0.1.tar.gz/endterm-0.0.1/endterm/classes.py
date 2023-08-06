import py7zr
import os
import xlrd
from openpyxl import load_workbook
import datetime as dt
from abc import ABC
import pandas as pd
from sys import platform

ZIP = [
    'ТЗА 2010.7z',
    'ТЗА 2011.7z',
    'ТЗА 2012.7z',
    'ТЗА 2013.7z',
    'ТЗА 2014.7z',
    'ТЗА 2015.7z',
    'ТЗА 2016.7z',
    'ТЗА 2017.7z',
    'ТЗА 2018.7z',
    'ТЗА 2019.7z',
    'ТЗА 2020.7z'
    ]


BASE_DIR = os.path.dirname(os.path.realpath(__file__))

WORK_DIR = os.getcwd()

class ZipExtract():


    def __init__(self):
        try:
            if platform == "linux" or platform == "linux2":
                TEMP = '/temp'
                for i in ZIP:
                    archive = py7zr.SevenZipFile(i, mode='r')
                    archive.extractall(path=WORK_DIR+TEMP)
                    archive.close()    
            elif platform == "darwin":
                TEMP = '/temp'
                for i in ZIP:
                    archive = py7zr.SevenZipFile(i, mode='r')
                    archive.extractall(path=WORK_DIR+TEMP)
                    archive.close()
            elif platform == "win32":
                TEMP = '\\temp'
                for i in ZIP:
                    archive = py7zr.SevenZipFile(i, mode='r')
                    archive.extractall(path=WORK_DIR+TEMP)
                    archive.close()
        except FileNotFoundError:
            raise FileNotFoundError("УСТОНОВИ ВСЕ АРХИВЫ В ТУЖЕ ПАПКУ В КОТОРОЙ РАБОТАЕШЬ")



class ReaderFactory(ABC):

    def extract_from_file(self):
        pass


class XLSReader(ReaderFactory):

    def __init__(self,year,name,path):
        self.year = year
        self.name = name
        self.xls = xlrd.open_workbook(path)
        self.xls_worksheet = None   
        self.dataframe = None

    def select_worksheet(self,number):
        if number > 3:
            raise ValueError('Number must be from 1 to 4')
        self.xls_worksheet = self.xls.sheets()[number]
        return self.xls_worksheet


    def extract_one_worksheet(self):
        cells_date = self.xls_worksheet.col_slice(0, 11, 200)
        time_list = []
        for cell in cells_date:
            time = xlrd.xldate_as_tuple(cell.value, 0)
            time_list.append(dt.date(time[0],time[1],time[2]))
        def to_int(cells):
            array = []
            for i in cells:
                if type(i) == str:
                    array.append(None)
                else:
                    array.append(int(i))
            return array
        def find_none(cells):
            array = []
            for i in cells:
                if type(i) == str:
                    array.append(None)
                else:
                    array.append(i)
            return array
        cells_2 = self.xls_worksheet.col_values(1,11,200)
        cells_2 = to_int(cells_2) # срок
        cells_3 = self.xls_worksheet.col_values(2,11,200) 
        cells_3 = find_none(cells_3) # взвеш.в-ва
        cells_4 = self.xls_worksheet.col_values(3,11,200) # диоксид серды
        cells_4 = find_none(cells_4)
        cells_5 = self.xls_worksheet.col_values(4,11,200) # сульфаты
        cells_5 = find_none(cells_5)
        cells_6 = self.xls_worksheet.col_values(5,11,200) # углерода оксид
        cells_6 = to_int(cells_6)
        cells_7 = self.xls_worksheet.col_values(6,11,200) # азот диоксида
        cells_7 = find_none(cells_7)
        cells_8 = self.xls_worksheet.col_values(7,11,200) # фтористый водород
        cells_8 = find_none(cells_8)
        cells_9 = self.xls_worksheet.col_values(8,11,200) # температура
        cells_9 = find_none(cells_9)
        cells_10 = self.xls_worksheet.col_values(9,11,200) # напр.града
        cells_10 = to_int(cells_10)
        cells_11 = self.xls_worksheet.col_values(10,11,200) # скор м.с
        cells_11 = to_int(cells_11)
        cells_12 = self.xls_worksheet.col_values(11,11,200) # атм.явление
        cells_12 = to_int(cells_12)
        data = {
            "Дата":time_list,
            "Срок":cells_2,
            "Взвеш.в-ва":cells_3,
            "Диоксид серы":cells_4,
            "Сульфаты":cells_5,
            "Углерода оксид":cells_6,
            "Азот диоксида":cells_7,
            "Фтористый водород":cells_8,
            "Температура":cells_9,
            "Напр.града":cells_10,
            "Скор м.с":cells_11,
            "Атм.явление":cells_12,
        }
        return data

    def extract_all_worksheet(self):
        array = []
        self.select_worksheet(0)
        array.append(self.extract_one_worksheet())
        self.select_worksheet(1)
        array.append(self.extract_one_worksheet())
        self.select_worksheet(2)
        array.append(self.extract_one_worksheet())
        self.select_worksheet(3)
        array.append(self.extract_one_worksheet())
        return array

    def to_dataframe(self):
        return pd.DataFrame(self.extract_one_worksheet())
    
    def create_dataframe(self):
        self.dataframe = pd.DataFrame(self.extract_one_worksheet())

    def __str__(self):
        return "<{} - {}>".format(self.year,self.xls)

    def __repr__(self):
        return "<{} - {}>".format(self.year,self.xls)



class XLSXReader(ReaderFactory):

    def __init__(self,year,name,path):
        self.year = year
        self.name = name
        self.xlsx = load_workbook(path)
        self.xlsx_worksheet = None   
        self.dataframe = None

    def select_worksheet(self,number):
        if number > 3:
            raise ValueError('Number must be from 1 to 4')
        temporary = self.xlsx.sheetnames[number]
        self.xlsx_worksheet = self.xlsx[temporary]
        return self.xlsx_worksheet

    def loop_iterate(self,word):
        array = [] 
        def iterate():
            for row in self.xlsx_worksheet[word+'10':word+str(len(list(self.xlsx_worksheet.rows)))]:
                for cell in row:
                    yield cell.value
        for i in iterate():
            array.append(i)
        return array
        

    def extract_one_worksheet(self):
        cells1 = self.loop_iterate('A')
        cells2 = self.loop_iterate('B')
        cells3 = self.loop_iterate('C')
        cells4 = self.loop_iterate('D')
        cells5 = self.loop_iterate('E')
        cells6 = self.loop_iterate('F')
        cells7 = self.loop_iterate('G')
        cells8 = self.loop_iterate('H')
        cells9 = self.loop_iterate('I')
        cells10 = self.loop_iterate('J')
        cells11 = self.loop_iterate('K')
        cells12 = self.loop_iterate('L')
        data = {
            "Дата":cells1,
            "Срок":cells2,
            "Взвеш.в-ва":cells3,
            "Диоксид серы":cells4,
            "Сульфаты":cells5,
            "Углерода оксид":cells6,
            "Азот диоксида":cells7,
            "Фтористый водород":cells8,
            "Температура":cells9,
            "Напр.града":cells10,
            "Скор м.с":cells11,
            "Атм.явление":cells12,
        }
        return data

    def to_dataframe(self):
        return pd.DataFrame(self.extract_one_worksheet())[2:]
    
    def reset_dataframe(self):
        return self.to_dataframe().reset_index(drop=True)
    
    def create_dataframe(self):
        self.dataframe = self.reset_dataframe()
    

    def __str__(self):
        return "<{} - {}>".format(self.year,self.xlsx)

    def __repr__(self):
        return "<{} - {}>".format(self.year,self.xlsx)

