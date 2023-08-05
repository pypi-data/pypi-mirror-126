# -*- encoding: utf-8 -*-
'''cli_clslq_notion

Notion client in python, notion-py is required

Usage: clslq notion [OPTIONS]

[openpyxl](https://openpyxl-chinese-docs.readthedocs.io/zh_CN/latest/tutorial.html)

[pandas](https://www.pypandas.cn/)

Notion template on https://airy-skiff-4d0.notion.site/04143158def3413fb58e7dae4b2d9aff

'''

import click
import datetime
import string
import time
import json
import os
import requests
import base64
import re
import traceback
import pandas
import calendar
# Support email
import smtplib
from io import BytesIO
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill

from .clslq_config import ClslqConfigUnique
from .clslq_log import ClslqLogger
from notion_client import Client

from .templates import weekreport
from .templates import monthreport

clslog = ClslqLogger().log


class Report(object):
    """Base class of Report

    Args:
        object (wbname): Workbook name, generated document file name
    """
    def __init__(self, wbname):
        self.wb = Workbook()
        self.sht = self.wb.active
        self.default_border = Border(left=Side(border_style='thin',
                                               color='000000'),
                                     right=Side(border_style='thin',
                                                color='000000'),
                                     top=Side(border_style='thin',
                                              color='000000'),
                                     bottom=Side(border_style='thin',
                                                 color='000000'))
        self.wbname = wbname
        self.datetime_now = datetime.datetime.now()

    @property
    def now(self):
        return self.datetime_now

    def render_html_without_inline_css(self, title, df):
        """Render html form template, use pandas

        Note that some of the email display methods only support inline-css style,
        This method does not use any inline-css for table headers and rows.

        Deprecated

        Args:
            title (str): Title of html and email subject
            df (object): Pandas DataFrame object
        """
        clslog.info("Render html for {}".format(title))
        with open(self.wbname + '.html', encoding='utf-8', mode='w') as f:

            task = df[df[u'分类'] != u'工作计划']
            plan = df[df[u'分类'] == u'工作计划']

            t = string.Template(weekreport.wr_template)
            f.write(
                t.safe_substitute({
                    "title":
                    title,
                    "table":
                    task.to_html(classes='tablestyle', index=False, na_rep=""),
                    "plan":
                    plan.to_html(classes='tablestyle', index=False, na_rep="")
                }))

    def send_study_email(self, config, title, email_html):
        """Send report email to receivers defined in .clslq.json

        Args:
            config (dict): Loaded json object from .clslq.json
            title (str): Unicode string as email subject
        """
        email = config.get('email')
        smtpserver = email['sender']['smtpserver']
        user = email['sender']['user']
        pwd = email['sender']['pwd']
        receivers = email['study_receivers']
        clslog.info("{} {} {}".format(smtpserver, user, title))

        msg = MIMEMultipart()
        msg['From'] = "{}".format(user)
        msg['To'] = ",".join(receivers)
        msg['Subject'] = Header(title, 'utf-8')
        msg.attach(MIMEText(email_html, 'html', 'utf-8'))
        try:
            smtp = smtplib.SMTP()
            smtp.connect(smtpserver)
            smtp.login(user, pwd)
            smtp.sendmail(user, receivers, msg.as_string())
            clslog.info('Email sent to {} done'.format(receivers))
            smtp.quit()
            smtp.close()
        except Exception as e:
            clslog.critical("Exception @{}: {}".format(
                e.__traceback__.tb_lineno, e))

    def send_email(self, config, title):
        """Send report email to receivers defined in .clslq.json

        Args:
            config (dict): Loaded json object from .clslq.json
            title (str): Unicode string as email subject
        """
        email = config.get('email')
        smtpserver = email['sender']['smtpserver']
        user = email['sender']['user']
        pwd = email['sender']['pwd']
        receivers = email['receivers']
        clslog.info("{} {} {}".format(smtpserver, user, title))

        msg = MIMEMultipart()
        msg['From'] = "{}".format(user)
        msg['To'] = ",".join(receivers)
        msg['Subject'] = Header(title, 'utf-8')
        with open(self.wbname + '.html', "r", encoding='utf-8') as f:
            msg.attach(MIMEText(f.read(), 'html', 'utf-8'))
        try:
            smtp = smtplib.SMTP()
            smtp.connect(smtpserver)
            smtp.login(user, pwd)
            smtp.sendmail(user, receivers, msg.as_string())
            clslog.info('Email sent to {} done'.format(receivers))
            smtp.quit()
            smtp.close()
        except Exception as e:
            clslog.critical("Exception @{}: {}".format(
                e.__traceback__.tb_lineno, e))

    def remove_files(self):
        """Removes all generated files"""
        if os.path.exists(self.wbname + '.html'):
            os.remove(self.wbname + '.html')
        if os.path.exists(self.wbname + '.xlsx'):
            os.remove(self.wbname + '.xlsx')


class WeekReport(Report):
    def week_belongs(self, sdate, edate):
        s = sdate - datetime.timedelta(days=1)
        e = edate + datetime.timedelta(days=1)
        clslog.info("[{} {}] {}".format(s, e, self.datetime_now))
        if self.datetime_now >= (sdate - datetime.timedelta(
                days=2)) and self.datetime_now <= (edate +
                                                   datetime.timedelta(days=2)):
            return True
        else:
            return False

    def set_worksheet_head(self, head):
        """Set sheet head

        Args:
            head (worksheet head): string
        """
        self.sht.merge_cells('A1:F1')

        self.sht['A1'] = u"本周工作情况" + head.replace('-', "")
        self.sht['A1'].alignment = Alignment(horizontal='center',
                                             vertical='center')
        self.sht['A1'].font = Font(color=colors.BLACK, b=True, size=14)
        self.sht['A1'].fill = PatternFill("solid", fgColor="00FF8080")
        self.sht.row_dimensions[1].height = 20

    def set_worksheet_title(self):
        u"""openpyxl 不支持按列设置样式
        """
        self.sht['A2'] = u"分类"
        self.sht['B2'] = u"事项"
        self.sht['C2'] = u"进展"
        self.sht['D2'] = u"问题"
        self.sht['E2'] = u"解决"
        self.sht['F2'] = u"评审、复盘、总结"
        for col in range(1, 7):
            self.sht.cell(column=col, row=2).font = Font(name=u'微软雅黑',
                                                         bold=True,
                                                         size=12)

    def content_parse_title(self, item):
        result = None
        try:
            title = item[u'名称']['title']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_state(self, item, row):
        result = ''
        try:
            for i in item[u'状态']['multi_select']:
                if i['name'] == u'进行中':
                    for col in range(1, 7):
                        self.sht.cell(column=col, row=row).fill = PatternFill(
                            "solid", fgColor="00CCFFCC")
                if i['name'] == u'遇问题':
                    for col in range(1, 7):
                        self.sht.cell(column=col, row=row).fill = PatternFill(
                            "solid", fgColor="00FFCC99")
                result = "{} {}".format(result, i['name'])
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_type(self, item, row):
        result = ''
        try:
            result = item[u'分类']['select']['name']
            if result == u'工作计划':
                self.sht.cell(column=6, row=row).value = u"下周工作计划"
                for col in range(1, 7):
                    self.sht.cell(column=col, row=row).fill = PatternFill(
                        "solid", fgColor="00CCFFCC")
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_richtext(self, item, text):
        """Parse Notion Column content

        Args:
            item (dict): Notion Column content
            text (str): Unicode string means column title

        Returns:
            str: Cell result
        """
        result = ''
        try:
            title = item[text]['rich_text']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result.replace('\n', ' ')

    def excel_worksheet_fill(self, item, row):
        self.sht['A' + str(row)] = self.content_parse_type(item, row)
        self.sht['B' + str(row)] = self.content_parse_title(item)
        self.sht['C' + str(row)] = self.content_parse_state(item, row)
        self.sht['D' + str(row)] = self.content_parse_richtext(item, u'问题')
        self.sht['E' + str(row)] = self.content_parse_richtext(item, u'解决方法')
        self.sht['F' + str(row)] = self.content_parse_richtext(
            item, u'评审、复盘、总结')

        for c in ('A', 'C'):
            self.sht.column_dimensions[c].width = 10
            for cell in self.sht[c]:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)
                cell.border = self.default_border
        for c in ('B', 'F', 'D', 'E'):
            self.sht.column_dimensions[c].width = 30
            for cell in self.sht[c]:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)
                cell.border = self.default_border

    def excel_worksheet_dump(self, wbname, database):
        # Change sheet name
        self.sht.title = 'WR'
        self.set_worksheet_head(wbname)
        self.set_worksheet_title()

        row = 3
        for node in database['results']:
            item = node['properties']
            self.excel_worksheet_fill(item, row)
            row = row + 1
        self.wb.save(filename=wbname + '.xlsx')
        self.wb.close()

    def pandas_df_fill(self, database):
        _type = []
        _title = []
        _state = []
        _problem = []
        _solve = []
        _conclusion = []
        for node in database['results']:
            item = node['properties']
            _type.append(self.content_parse_type(item, 0))
            _title.append(self.content_parse_title(item))
            _state.append(self.content_parse_state(item, 0))
            _problem.append(self.content_parse_richtext(item, u'问题'))
            _solve.append(self.content_parse_richtext(item, u'解决方法'))
            _conclusion.append(self.content_parse_richtext(item, u'评审、复盘、总结'))
        data = {
            u'分类': pandas.Series(_type, index=range(len(_type))),
            u'事项': pandas.Series(_title, index=range(len(_title))),
            u'进展': pandas.Series(_state, index=range(len(_state))),
            u'问题': pandas.Series(_problem, index=range(len(_problem))),
            u'解决': pandas.Series(_solve, index=range(len(_solve))),
            u'评审复盘总结备注': pandas.Series(_conclusion,
                                       index=range(len(_conclusion)))
        }
        self.df = pandas.DataFrame(data)
        return self.df

    def render_html(self, clsconfig, title, database):
        """Render html form template

        Note that some of the email display methods only support inline-css style,
        This method support inline-css for table headers and rows.

        Args:
            title (str): Title of html and email subject
            df (object): Pandas DataFrame object
        """
        clslog.info("Render[Inline-CSS] html for {}".format(title))
        with open(self.wbname + '.html', encoding='utf-8', mode='w') as f:
            table = str('')
            summary = str('')
            plan = str('')
            i = 0
            j = 0

            task_paln_template = """
                <tr height="19" style="height:14.0pt;background: {color}">
                    <td style="border: 0.5pt solid #cfcfcf; vertical-align: middle; text-align: center;">{type}</td>
                    <td style="border: 0.5pt solid #cfcfcf;">{title}</td>
                    <td style="border: 0.5pt solid #cfcfcf; vertical-align: middle; text-align: center;">{state}</td>
                    <td style="border: 0.5pt solid #cfcfcf;">{problem}</td>
                    <td style="border: 0.5pt solid #cfcfcf;">{solve}</td>
                </tr>
            """
            summary_template = """
                <tr>
                    <td style="padding: 10px; background-color: rgba(204, 204, 204, 0.1)">
                    <span style="font-size: 16px; color: #81e4c3">●</span>&nbsp;
                    <span>
                        <span style="border-bottom: 1px dashed rgb(204, 204, 204); position: relative;">{summary}</span>
                    </span>
                    </td>
                </tr>
            """
            for node in database['results']:
                item = node['properties']

                def bgcolor(i):
                    return '#F7F7F7' if i % 2 == 0 else '#fff'

                if self.content_parse_type(item, 0) == u'工作计划':
                    plan += task_paln_template.format(
                        **{
                            'type': str(i + 1),
                            'title': self.content_parse_title(item),
                            'state': "-",
                            'problem': self.content_parse_richtext(
                                item, u'问题'),
                            'solve': self.content_parse_richtext(
                                item, u'解决方法'),
                            'color': bgcolor(i)
                        })
                    i = i + 1
                else:
                    table += task_paln_template.format(
                        **{
                            'type': self.content_parse_type(item, 0),
                            'title': self.content_parse_title(item),
                            'state': self.content_parse_state(item, 0),
                            'problem': self.content_parse_richtext(
                                item, u'问题'),
                            'solve': self.content_parse_richtext(
                                item, u'解决方法'),
                            'color': bgcolor(j)
                        })
                    j = j + 1
                summarize_item = self.content_parse_richtext(item, u'评审、复盘、总结')
                if len(summarize_item):
                    summary += summary_template.format(
                        **{'summary': summarize_item})
            html = string.Template(weekreport.wr_template)

            f.write(
                html.safe_substitute({
                    "title": title,
                    "table": table,
                    "plan": plan,
                    "user": clsconfig.get('user'),
                    "department": clsconfig.get('department'),
                    "summarize": summary
                }))


class MonthReport(Report):
    def belongs(self, date):
        if date > self._mdaystart and date <= self._mdayend:
            return True
        else:
            return False

    def week_belongs(self, sdate, edate):
        if sdate >= (self._mdaystart -
                     datetime.timedelta(days=1)) and edate <= (
                         self._mdayend + datetime.timedelta(days=1)):
            return True
        else:
            return False

    def init(self, title, user, department, force):

        nowdate = self.datetime_now
        if nowdate.month == 1:
            lastmonth = 12
        else:
            lastmonth = nowdate.month - 1
        self._mdaystart = nowdate - \
            datetime.timedelta(
                days=(nowdate.day + calendar.mdays[lastmonth]))
        if force:
            self._mdayend = nowdate
        else:
            self._mdayend = nowdate - datetime.timedelta(days=nowdate.day)

        clslog.critical("==月报检索{}==>{}周报及读书学习笔记==".format(
            self._mdaystart.strftime("%Y%m%d"),
            self._mdayend.strftime("%Y%m%d")))

        # 当月发上月报告
        mtitle = "{}({}{:02})".format(title, nowdate.year, lastmonth)
        click.secho("{}".format(mtitle), fg='blue')

        self._title = mtitle

        self._book = str('')
        self._book_content = str('')
        self._study_list = str('')
        self._main_target = str('')
        self._team_target = str('')
        self._technology = str('')
        self._patent = str('')
        self._review = str('')
        self._tech_issues = str('')
        self._maintainance = str('')
        self._duties = str('')
        self._programming_tasks = str('')
        self._reading_share = str('')
        _template = """
            <li style="list-style-type: upper-roman;">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span></li>
        """
        self._reading_share += _template.format(**{
            'content': u"项目例会中结合技术管理、结构思考等方面书中所学",
            'solve': u"会议提议与解决着重高效"
        })
        self._reading_share += _template.format(
            **{
                'content': u"日常工作结合技术分享精神",
                'solve': u"利用自身高效整理的特点，将团队技术栈相关的有用信息分享给团队开发人员"
            })
        self._directions = str('')

        self._user = user
        self._department = department

    @property
    def mtitle(self):
        return self._title

    @mtitle.setter
    def mtitle(self, xtitle):
        if not isinstance(xtitle, str):
            raise ValueError("mtitle must be an string")
        else:
            self._title = xtitle

    def content_parse_title(self, item):
        result = None
        try:
            title = item[u'名称']['title']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_state(self, item):
        result = ''
        try:
            for i in item[u'状态']['multi_select']:
                result = "{} {}".format(result, i['name'])
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_type(self, item):
        result = ''
        try:
            result = item[u'分类']['select']['name']
            if result == u'工作计划':
                return ''
        except Exception as e:
            pass
        finally:
            return result

    def content_parse_richtext(self, item, text):
        """Parse Notion Column content

        Args:
            item (dict): Notion Column content
            text (str): Unicode string means column title

        Returns:
            str: Cell result
        """
        result = ''
        try:
            title = item[text]['rich_text']
            for i in title:
                result = i['plain_text']
        except Exception as e:
            pass
        finally:
            return result.replace('\n', ' ')

    def block_parse_table_cell_properties(self, properties, key):
        _type = properties[key]['type']
        _result = ''

        def date_valid(x):
            return x if x is not None else ""

        def num_valid(x):
            return str(x) if x is not None else ""

        try:
            if _type == 'rich_text' or _type == 'title':
                for i in properties[key][_type]:
                    _result += i['plain_text']

            elif _type == 'multi_select':
                for i in properties[key][_type]:
                    _result += """<span style="color:{color};">{content}</span>""".format(
                        **{
                            'color': i['color'],
                            'content': i['name']
                        })
            elif _type == 'select':
                i = properties[key][_type]
                _result += """<span style="color:{color};">{content}</span>""".format(
                    **{
                        'color': i['color'],
                        'content': i['name']
                    })
            elif _type == 'url':
                _result += """<a href="{url}" target="_blank">{text}</a>""".format(
                    **{
                        'url': properties[key][_type],
                        'text': properties[key][_type]
                    })
            elif _type == 'date':
                _result += """{s} {e}""".format(
                    **{
                        's': date_valid(properties[key][_type]['start']),
                        'e': date_valid(properties[key][_type]['end'])
                    })
            elif _type == 'number':
                _result += num_valid(properties[key][_type])
            elif _type == 'files':
                for f in properties[key][_type]:

                    if 'file' in f:
                        url = f['file']['url']
                    elif 'external' in f:
                        url = f['external']['url']

                    img_template = """
                        <img style="width: 200px;" src="{img}"></img>
                    """
                    imgsrc = self.img_url_to_base64(url)
                    clslog.info(imgsrc[0:32] + '...')
                    if imgsrc:
                        _result += img_template.format(**{'img': imgsrc})
                    else:
                        _result += img_template.format(**{'img': url})
            else:
                clslog.warning("Unsupported: {}".format(properties[key]))
        except Exception as e:
            clslog.critical("Exception @{}: {}".format(
                e.__traceback__.tb_lineno, e))

        return _result

    def block_parse_table(self, client, b):
        database = client.databases.query(b['id'])
        table = str(
            '<table border="0" cellpadding="0" cellspacing="0" style="border-collapse: collapse; color: #3d3b4f; background-color: #fff; border: 1px solid #cfcfcf; box-shadow: 0 0px 6px rgba(0, 0, 0, 0.1); margin-bottom: 20px; font-size:13px;">'
        )
        heads = []
        if len(database['results']):
            table += '<thead>'
            for h in database['results'][0]['properties']:
                heads.append(h)
            for hi in reversed(heads):
                table += """<th height="19" width="107"
                          style="border: 0.5pt solid #cfcfcf; width: 107pt; height: 14pt; padding-top: 1px; padding-right: 1px; padding-left: 1px; font-size: 11pt; font-family: 宋体; vertical-align: middle; text-align: center;">{h}</th>""".format(
                    **{'h': hi})
            table += '</thead>'
            table += '<tbody>'
            for i in database['results']:
                p = i['properties']
                table += '<tr height="19" style="height:14.0pt;background-color:rgb(253, 233, 217)">'
                for hi in reversed(heads):
                    table += """<td height="19" width="107"
                          style="border: 0.5pt solid #cfcfcf; width: 107pt; height: 14pt; padding-top: 1px; padding-right: 1px; padding-left: 1px; font-size: 11pt; font-family: 宋体; vertical-align: middle; text-align: center;">{v}</td>""".format(
                        **{'v': self.block_parse_table_cell_properties(p, hi)})
                table += '</tr>'
            table += '</tbody>'
        table += str('</table>')
        return table

    def block_common_types(self, client, b):
        children = str('')

        img_template = """
            <div style="width:100%;border: 1px dashed #ccc;
                vertical-align: middle;text-align: center; 
                box-shadow: 0px 0px 5px 5px rgba(10,10,10,0.2);
                -moz-box-shadow: 0px 0px 5px 5px rgba(10,10,10,0.2);
                -webkit-box-shadow: 0px 0px 5px 5px rgba(10,10,10,0.2);" >
                <img style="width: 80%; margin-top: 50px; margin-bottom: 50px;" src="{img}"></img>
            </div>
        """
        try:
            if b['type'] == "image":
                try:
                    if 'file' in b['image']:
                        url = b['image']['file']['url']
                    elif 'external' in b['image']:
                        url = b['image']['external']['url']

                    imgsrc = self.img_url_to_base64(url)
                    clslog.info(imgsrc[0:32] + '...')
                    if imgsrc:
                        children += img_template.format(**{'img': imgsrc})
                    else:
                        children += img_template.format(**{'img': url})
                except Exception as e:
                    clslog.critical("Exception @{}: {}".format(
                        e.__traceback__.tb_lineno, e))
                    clslog.info(b)

            elif b['type'] == "paragraph":
                text = str('')
                for t in b['paragraph']['text']:
                    text += t['plain_text']
                children += """<p>{p}</p>""".format(**{'p': text.strip()})

            elif b['type'] == 'heading_1':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<h1>{head}</h1>""".format(
                    **{'head': text.strip()})
            elif b['type'] == 'heading_2':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<h2>{head}</h2>""".format(
                    **{'head': text.strip()})
            elif b['type'] == 'heading_3':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<h3>{head}</h3>""".format(
                    **{'head': text.strip()})
            elif b['type'] == 'heading_4':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<h4>{head}</h4>""".format(
                    **{'head': text.strip()})
            elif b['type'] == 'heading_5':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<h5>{head}</h5>""".format(
                    **{'head': text.strip()})
            elif b['type'] == 'heading_6':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<h6>{head}</h6>""".format(
                    **{'head': text.strip()})
            elif b['type'] == 'code':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<pre style="background-color: f5f5f5"><code>{code}</code></pre>""".format(
                    **{'code': text.strip()})
            elif b['type'] == 'bookmark':
                text = str('')
                for t in b[b['type']]['caption']:
                    text += t['plain_text']
                children += """<a href="{bookmark}" target="_blank">{caption}</a><br/>""".format(
                    **{
                        'bookmark': b[b['type']]['url'],
                        'caption': text.strip()
                    })
            elif b['type'] == 'child_database':
                children += self.block_parse_table(client, b)
            elif b['type'] == 'callout':
                text = str('')
                for t in b[b['type']]['text']:
                    text += t['plain_text']
                children += """<pre style="background-color: f5f5f5"><code>{icon}{code}</code></pre>""".format(
                    **{
                        'code': text.strip(),
                        'icon': b[b['type']]['icon']['emoji']
                    })
            else:
                clslog.warning("An unsupported type was received")
                clslog.info(b)
        except Exception as e:
            clslog.critical("Exception @{}: {}".format(
                e.__traceback__.tb_lineno, e))
        return children

    def block_list_children(self, client, block, tag, level=0):
        """Recursive function"""
        bs = client.blocks.children.list(block['id'])
        children = str('')
        if tag == 'ol':
            if level % 4 == 0:
                children = "<" + tag + " type=\"1\">"
            elif level % 4 == 1:
                children = "<" + tag + " type=\"A\">"
            elif level % 4 == 2:
                children = "<" + tag + " type=\"a\">"
            elif level % 4 == 3:
                children = "<" + tag + " type=\"I\">"
            else:
                children = "<" + tag + " type=\"i\">"
        elif tag == 'ul':
            if level % 4 == 0:
                children = "<" + tag + " style=\"list-style-type:\1F44D\">"
            elif level % 4 == 1:
                children = "<" + tag + " style=\"list-style-type:\1F44D\">"
            elif level % 4 == 2:
                children = "<" + tag + " style=\"list-style-type:disc\">"
            elif level % 4 == 3:
                children = "<" + tag + " style=\"list-style-type:\1F44D\">"
            else:
                children = "<" + tag + " style=\"none\">"

        for i in range(len(bs['results'])):
            b = bs['results'][i]

            if b['type'] == 'numbered_list_item' or b[
                    'type'] == 'bulleted_list_item':

                plain_text = str('')
                for t in b[b['type']]['text']:
                    plain_text += t['plain_text']
                if b['has_children']:
                    level += 1
                    new_children = str('')
                    if b['type'] == 'bulleted_list_item':
                        new_children += self.block_list_children(
                            client, b, 'ul', level)
                    elif b['type'] == 'numbered_list_item':
                        new_children += self.block_list_children(
                            client, b, 'ol', level)
                    children += """<li>{item}{children}</li>""".format(
                        **{
                            'item': plain_text,
                            'children': new_children
                        })
                else:
                    children += """<li>{item}</li>""".format(
                        **{'item': plain_text})
            else:
                children += self.block_common_types(client, b)

        children += "</" + tag + ">"
        return children

    def render_block_list(self, client, block, tag, level):
        """Recurisve <ol><ul><li> tags management"""
        plain_text = str('')
        for t in block[block['type']]['text']:
            plain_text += t['plain_text']
        if block['has_children']:
            level += 1
            last = """<li>{item}{children}</li>""".format(
                **{
                    'item': plain_text,
                    'children': self.block_list_children(
                        client, block, tag, level)
                })
        else:
            last = """<li>{item}</li>""".format(**{'item': plain_text})
        # clslog.info(last)
        return last

    def img_url_to_base64(self, url):
        try:
            imgtype = os.path.splitext(url)[1].split('?')[0].replace('.', '/')
            response = requests.get(url)
            imgbase64 = base64.b64encode(BytesIO(response.content).read())
            return "data:image{};base64,".format(imgtype) + imgbase64.decode(
                "utf-8")
        except Exception as e:
            clslog.critical("Exception @{}: {}".format(
                e.__traceback__.tb_lineno, e))
            return None

    def render_block_items(self, client, page):
        page_html_tags = str('')
        p_template = """
            <p style="text-indent: 2em;">{p}</p>
        """
        img_template = """
            <div style="width:auto;border: 1px dashed #ccc;display: table-cell;
                vertical-align: middle;text-align: center; 
                box-shadow: 0px 0px 5px 5px rgba(10,10,10,0.2);
                -moz-box-shadow: 0px 0px 5px 5px rgba(10,10,10,0.2);
                -webkit-box-shadow: 0px 0px 5px 5px rgba(10,10,10,0.2);" >
                <img style="width: 80%;" src="{img}"></img>
            </div>
        """

        page_id = page['id']
        page_blocks = client.blocks.retrieve(page_id)
        page_content = client.blocks.children.list(page_blocks['id'])
        # clslog.info(json.dumps(page_content))
        """Supported blocks:
        paragraph, numbered_list_item, bulleted_list_item
        emoji icons
        """
        bulleted_list = str('')
        numbered_list = str('')
        last_type = None
        for i in range(len(page_content['results'])):
            b = page_content['results'][i]
            # End of <ol>
            if (last_type == 'numbered_list_item'
                    and b['type'] != 'numbered_list_item') or i == len(
                        page_content['results']) - 1:
                page_html_tags += """<ol type="1">{list}</ol>""".format(
                    **{'list': numbered_list})
                numbered_list = str('')
            # End of <ul>
            if (last_type == 'bulleted_list_item'
                    and b['type'] != 'bulleted_list_item') or i == len(
                        page_content['results']) - 1:
                page_html_tags += """<ul style="list-style-type:disc">{list}</ul>""".format(
                    **{'list': bulleted_list})
                bulleted_list = str('')
            if b['type'] == 'numbered_list_item':
                numbered_list += self.render_block_list(client, b, 'ol', 0)
            elif b['type'] == 'bulleted_list_item':
                bulleted_list += self.render_block_list(client, b, 'ul', 0)

            elif b['type'] == 'paragraph':
                text = str('')
                for t in b['paragraph']['text']:
                    text += t['plain_text']
                page_html_tags += p_template.format(**{'p': text.strip()})
            elif b['type'] == 'image' \
                or b['type'] == 'heading_1' \
                or b['type'] == 'heading_2' \
                or b['type'] == 'heading_3' \
                or b['type'] == 'heading_4' \
                or b['type'] == 'heading_5' \
                or b['type'] == 'heading_6' \
                or b['type'] == 'code' \
                    or b['type'] == 'bookmark':
                page_html_tags += self.block_common_types(client, b)
            else:
                clslog.warning("Type:{}".format(b['type']))
                bs = client.blocks.retrieve(b['id'])
                clslog.info(
                    "{} Not supported by current Notion API, more info please visit {}"
                    .format(
                        bs,
                        "https://developers.notion.com/docs/working-with-page-content#modeling-content-as-blocks"
                    ))

            last_type = b['type']
        return page_html_tags

    def render_reading_books(self, client, database):
        for page in database['results']:
            item = page['properties']
            date = datetime.datetime.fromisoformat(
                item[u'收录日期']['date']['start'])
            """Skip useless databases
            """
            if self.belongs(date):
                self._book += self.content_parse_title(item)
                self._book_content += self.render_block_items(client, page)

    def render_study_note(self, client, database):

        for page in database['results']:
            item = page['properties']
            date = datetime.datetime.fromisoformat(
                item[u'收录日期']['date']['start'])
            """Skip useless databases
            """
            if self.belongs(date):
                title_url = item[u'链接']['url']
                self._study_list += """<h1>{title}</h1><br/>""".format(
                    **{'title': self.content_parse_title(item)})
                if title_url:
                    self._study_list += """<a href="{url}" target="_blank"><p>{title}</p></a><br/>""".format(
                        **{
                            'title': "扩展阅读",
                            'url': title_url
                        })

                self._study_list += self.render_block_items(client, page)
        # clslog.info(self._study_list)

    def render_maintarget(self, database):
        for node in database['results']:
            item = node['properties']
            self._main_target += """<li style="list-style-type: demical;">{p}</li>""".format(
                **{'p': item[u'目标']['title'][0]['plain_text']})

    def render_teamtarget(self, database):
        for node in database['results']:
            item = node['properties']
            self._team_target += """<li style="list-style-type: demical;">{p}</li>""".format(
                **{'p': item[u'目标']['title'][0]['plain_text']})

    def render_technology(self, database):
        _template = """
            <li style="list-style-type: upper-roman; ">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'方案输出' or _type == u'技术预研':
                self._technology += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_review(self, database):
        _template = """
            <li style="list-style-type: upper-roman; ">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'代码评审' or _type == u'配置管理':
                self._review += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_maintainance(self, database):
        _template = """
            <li style="list-style-type: upper-roman; ">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:blue;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'项目维护' or _type == u'IT运维':
                self._maintainance += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_patent(self, database):
        _template = """
            <li style="list-style-type: upper-roman;">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'知识产权建设':
                self._patent += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_techissues(self, database):
        _template = """
            <li style="list-style-type: upper-roman; ">
            <span style="font-weight: bold;">{content}</span>&nbsp;
            <span style="color:red;">{problem}</span>&nbsp;
            <span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'项目支撑' or _type == u'技术问题指导':
                self._tech_issues += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'problem': self.content_parse_richtext(item, u'问题'),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_duties(self, database):
        _template = """
            <li style="list-style-type: upper-roman;">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'技术管理' or _type == u'内部支撑':
                self._duties += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_programming_work(self, database):
        _template = """
            <li style="list-style-type: upper-roman;">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'产品开发':
                self._programming_tasks += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_reading_share(self, database):
        _template = """
            <li style="list-style-type: upper-roman;">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """
        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'总结输出' or _type == u'会议记录' or _type == u'复盘分享':
                self._reading_share += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def render_directions(self, database):
        _template = """
            <li style="list-style-type: upper-roman;">
            <span style="font-weight: bold;">{content}</span>
            &nbsp;<span style="color:green;">{solve}</span>
            &nbsp;<span style="color:blue;">{summary}</span></li>
        """

        for node in database['results']:
            item = node['properties']
            _type = self.content_parse_type(item)
            if _type == u'技术分享' or _type == u'好书分享' or _type == u'好文章分享' or _type == u'工作指导':
                self._directions += _template.format(
                    **{
                        'content': self.content_parse_title(item),
                        'summary': self.content_parse_richtext(
                            item, u'评审、复盘、总结'),
                        'solve': self.content_parse_richtext(item, u'解决方法')
                    })

    def parse_week_tasks(self, database, weekdatestr):
        self.render_technology(database)
        self.render_patent(database)
        self.render_review(database)
        self.render_duties(database)
        self.render_techissues(database)
        self.render_maintainance(database)
        self.render_programming_work(database)
        self.render_reading_share(database)
        self.render_directions(database)

    def render_itor_database(self, client, dbid, title, database, force):
        try:
            for t in title:
                # Use BT-Panel timer task to trigger
                if not force:
                    if not (self.datetime_now.day >= 1
                            and self.datetime_now.day <= 6):
                        clslog.warning(
                            "Month report will not trigger except the day 1")
                        break

                plain_text = t['plain_text'].strip().replace(' → ', '~')
                """Reading list"""
                if plain_text == u"读书如斯":
                    self.render_reading_books(client, database)
                    break
                if plain_text == u"网络万象":
                    self.render_study_note(client, database)
                    break
                if plain_text == u"公司整体目标":
                    self.render_maintarget(database)
                    break
                if plain_text == u"团队技术方向和目标":
                    self.render_teamtarget(database)
                    break
                """Week report parse"""
                plain_text_head = t['plain_text'][0:3]  # Filter for week tasks
                value = re.compile(r'^[0-9][0-9][0-9]$')
                if t['type'] == 'mention':
                    sdate = datetime.datetime.fromisoformat(
                        t['mention']['date']['start'])
                    edate = datetime.datetime.fromisoformat(
                        t['mention']['date']['end'])
                    weekdatestr = """{s}===>{e}""".format(
                        **{
                            's': sdate.strftime("%Y-%m-%d"),
                            'e': edate.strftime("%Y-%m-%d")
                        })
                    """Skip useless databases
                    """
                    if sdate == edate:
                        break
                    if value.match(plain_text_head):
                        if self.week_belongs(sdate, edate):
                            click.secho("解析{}工作任务".format(weekdatestr),
                                        fg='blue')
                            self.parse_week_tasks(database, weekdatestr)
                        else:
                            clslog.info(
                                "Week {} report N/A".format(weekdatestr))
                            break
        except Exception as e:
            clslog.critical("Exception @{}: {}".format(
                e.__traceback__.tb_lineno, e))
            traceback.print_exc(e)

    def render_html(self, title):
        """Render html form template

        Note that some of the email display methods only support inline-css style,
        This method support inline-css for table headers and rows.

        Args:
            title (str): Title of html and email subject
        """
        clslog.info("Render[Inline-CSS] html for {}".format(title))

        with open(self.wbname + '.html', encoding='utf-8', mode='w') as f:

            html = string.Template(monthreport.mr_template)

            f.write(
                html.safe_substitute({
                    "title": self._title,
                    "book": self._book,
                    "book_content": self._book_content,
                    "study": self._study_list,
                    "main_target": self._main_target,
                    "team_target": self._team_target,
                    "technology": self._technology,
                    "patent": self._patent,
                    "review": self._review,
                    "technology_issues": self._tech_issues,
                    "maintainance": self._maintainance,
                    "duties": self._duties,
                    "programming_work": self._programming_tasks,
                    "reading_share": self._reading_share,
                    "direction": self._directions,
                    "user": self._user,
                    "department": self._department
                }))

    def render_study_html(self, title):
        """Render html form template

        Note that some of the email display methods only support inline-css style,
        This method support inline-css for table headers and rows.

        Args:
            title (str): Title of html and email subject
        """
        clslog.info(
            "Render study content[Inline-CSS] html for {}".format(title))

        html = string.Template(monthreport.mstudy_template)

        return html.safe_substitute({
            "title": title,
            "book": self._book,
            "book_content": self._book_content,
            "study": self._study_list,
            "user": self._user,
            "department": self._department
        })


def cli_week(client, clsconfig, excel, remove, force, send):
    """Generate work report every week

    Args:
        client (object): Notion client instance
        clsconfig (dict): Config instance of .clslq.json
        excel (bool): Support dump excel or not
        remove (bool): Remove dumped files or not
        force (bool): Force generate or not
        send (bool): Send email or not
    """
    for i in client.search()['results']:
        if i['object'] == 'database':
            try:

                database = client.databases.query(i['id'])
                title = i['title']

                for t in title:
                    plain_text = t['plain_text'].strip().replace(' → ', '~')
                    plain_text_head = t['plain_text'][0:3]
                    value = re.compile(r'^[0-9]+[0-9]$')
                    # Handle valid report database only
                    if plain_text_head == 'WRT':
                        break
                    if value.match(plain_text_head) == None:
                        break
                    enddate = datetime.datetime.fromisoformat(
                        t['mention']['date']['end'])
                    startdate = datetime.datetime.fromisoformat(
                        t['mention']['date']['start'])

                    wrp = WeekReport(plain_text)

                    # Use BT-Panel timer task to trigger
                    if not force:
                        if wrp.now.weekday() != 5:  # 0~6 means Monday~Sunday
                            clslog.warning("Today is not Saturday")
                            break
                    if not wrp.week_belongs(startdate, enddate):
                        click.secho(
                            "[{}~{}] today:{}[weekday:{}] gap:{}days Week report expired"
                            .format(startdate.strftime("%Y-%m-%d"),
                                    enddate.strftime("%Y-%m-%d"),
                                    wrp.now.strftime("%Y-%m-%d"),
                                    wrp.now.weekday(),
                                    abs(startdate - wrp.now).days),
                            fg='green')
                        break
                    wtitle = "{}({})".format(clsconfig.get('wr_title_prefix'),
                                             plain_text)

                    if excel:
                        wrp.excel_worksheet_dump(plain_text, database)
                        df = pandas.read_excel(plain_text + '.xlsx',
                                               sheet_name='WR',
                                               header=1)
                    else:
                        df = wrp.pandas_df_fill(database)
                    pandas.set_option('colheader_justify', 'center')
                    df.style.hide_index()  # Hide index col

                    # with open("debug.json", "w") as file:
                    #     file.write(json.dumps(i))

                    wrp.render_html(clsconfig, wtitle, database)
                    if send:
                        wrp.send_email(clsconfig, wtitle)
                    if remove:
                        wrp.remove_files()

            except Exception as e:
                clslog.error(e)
                traceback.print_exc(e)


def cli_month(client, clsconfig, remove, force, send):
    """Generate work report every month

    Args:
        client (object): Notion client instance
        clsconfig (dict): Config instance of .clslq.json
        remove (bool): Remove dumped files or not
        force (bool): Force generate or not
        send (bool): Send email or not
    """
    mrp = MonthReport(clsconfig.get('mr_title_prefix'))
    mrp.init(clsconfig.get('mr_title_prefix'), clsconfig.get('user'),
             clsconfig.get('department'), force)
    # Use BT-Panel timer task to trigger
    if not force:
        if not (mrp.now.day >= 1 and mrp.now.day <= 5):
            clslog.warning(
                "Month report email only sent when mday [1~5] or in force mode"
            )
            return
    """Get week tasks, dump them into month table
    """
    for i in client.search()['results']:
        # Itor all database
        if i['object'] == 'database':
            try:
                database = client.databases.query(i['id'])
                title = i['title']
                mrp.render_itor_database(client, i['id'], title, database,
                                         force)

            except Exception as e:
                clslog.error(e)
                traceback.print_exc(e)
    try:
        mrp.render_html(mrp.mtitle)
        if force:
            study_title = "C(oncept)T(each)R(eview)S(implify)({}{:02})".format(
                mrp.now.year, mrp.now.month)
        else:
            study_title = "C(oncept)T(each)R(eview)S(implify)({}{:02})".format(
                mrp.now.year, mrp.now.month - 1)
        study_email = mrp.render_study_html(study_title)

        if send:
            mrp.send_email(clsconfig, mrp.mtitle)
            mrp.send_study_email(clsconfig, study_title, study_email)
        if remove:
            mrp.remove_files()
    except Exception as e:
        clslog.error(e)
        traceback.print_exc(e)


@click.option('--rtype',
              '-t',
              required=True,
              default='week',
              type=click.Choice(['week', 'month', 'all']),
              help='Choose a type to generate report')
@click.option('--excel',
              '-e',
              flag_value='GenerateExcel',
              default=False,
              help='Generate xlsx excel file or not')
@click.option('--remove',
              '-r',
              flag_value='RemoveFiles',
              help='Remove files or not')
@click.option(
    '--force',
    '-f',
    flag_value='force',
    default=False,
    help='Force generate right now, otherwise limited by valid datetime')
@click.option('--send',
              '-s',
              flag_value='send',
              default=False,
              help='Send email or not')
@click.option('--config',
              '-c',
              type=click.Path(exists=True),
              default='.clslq.json',
              help='CLSLQ config use {} as default.'.format('.clslq.json'))
@click.command(context_settings=dict(
    allow_extra_args=True,
    ignore_unknown_options=True,
),
               help="Notion Report Generator.")
def notion(rtype, config, excel, remove, force, send):

    clsconfig = ClslqConfigUnique(file=config)
    if clsconfig.get('secrets_from') is None:
        click.secho("Make sure notion secret code is valid in .clslq.json",
                    fg='red')
        return

    client = Client(auth=clsconfig.get('secrets_from'),
                    notion_version="2021-08-16")

    click.secho("Report type: {} Notion Accounts:".format(rtype), fg='yellow')
    for u in client.users.list()['results']:
        click.secho("{:8s} {}".format(u['name'], u['id']), fg='green')
    if 'week' == rtype:
        click.secho("Week report generator", fg='green')
        cli_week(client, clsconfig, excel, remove, force, send)
    elif 'month' == rtype:
        click.secho("Month report generator", fg='green')
        cli_month(client, clsconfig, remove, force, send)
    else:
        click.secho("All reports generator", fg='green')
        cli_week(client, clsconfig, excel, remove, force, send)
        cli_month(client, clsconfig, remove, force, send)
